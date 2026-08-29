"""
Data source loaders for SmartPOS AI.

Lets the app pull its product catalog from three places instead of only
the built-in mock data:
  1. Sample data (mock_data.py) — default, for demos.
  2. An uploaded Excel file.
  3. A database — either an uploaded SQLite file, or a live connection
     string (Postgres/MySQL/SQL Server/etc. via SQLAlchemy).

Every loader normalizes to the same schema the app already expects:
    product_id, name, category, price, cost, stock, reorder_level,
    expiry_date, margin_pct
so nothing downstream (Billing, Inventory, Dashboard) needs to change
based on where the data came from.
"""
import io
from datetime import datetime, timedelta

import pandas as pd

REQUIRED_COLUMNS = ["name", "category", "price", "cost", "stock", "reorder_level"]
OPTIONAL_COLUMNS = ["expiry_days", "gst_rate"]  # expiry_days: blank = non-perishable; gst_rate: blank = 0%


class DataValidationError(Exception):
    """Raised when uploaded/fetched data doesn't match the expected schema."""
    pass


def _normalize(df: pd.DataFrame) -> pd.DataFrame:
    """Validate columns, coerce types, and compute the derived fields the
    rest of the app relies on (product_id, expiry_date, margin_pct)."""
    df = df.copy()
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]

    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise DataValidationError(
            f"Missing required column(s): {', '.join(missing)}. "
            f"Required columns are: {', '.join(REQUIRED_COLUMNS)} "
            f"(optional: {', '.join(OPTIONAL_COLUMNS)})."
        )

    if "expiry_days" not in df.columns:
        df["expiry_days"] = None

    # Type coercion with clear errors instead of silent NaNs turning into bugs later
    numeric_cols = ["price", "cost", "stock", "reorder_level"]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    if df[numeric_cols].isna().any().any():
        bad_rows = df[df[numeric_cols].isna().any(axis=1)]
        raise DataValidationError(
            f"{len(bad_rows)} row(s) have non-numeric price/cost/stock/reorder_level. "
            f"Check rows: {bad_rows.index.tolist()[:10]}"
        )

    df["name"] = df["name"].astype(str).str.strip()
    df["category"] = df["category"].astype(str).str.strip()
    if df["name"].eq("").any() or df["name"].isna().any():
        raise DataValidationError("Every row needs a non-empty product name.")

    today = datetime.now()

    def to_expiry_date(v):
        if pd.isna(v) or v is None or str(v).strip() == "":
            return None
        try:
            return (today + timedelta(days=int(v))).date()
        except (ValueError, TypeError):
            return None

    df["expiry_date"] = df["expiry_days"].apply(to_expiry_date)
    df["margin_pct"] = ((df["price"] - df["cost"]) / df["price"].replace(0, pd.NA) * 100).round(1)
    df["margin_pct"] = df["margin_pct"].fillna(0)

    df = df.reset_index(drop=True)
    df["product_id"] = [f"P{i+1:03d}" for i in range(len(df))]

    cols = ["product_id", "name", "category", "price", "cost", "stock",
            "reorder_level", "expiry_date", "margin_pct"]
    return df[cols]


# --------------------------------------------------------------------------
# Excel
# --------------------------------------------------------------------------
def load_products_from_excel(uploaded_file) -> pd.DataFrame:
    """uploaded_file: a Streamlit UploadedFile (or any file-like object)."""
    try:
        raw = pd.read_excel(uploaded_file)
    except Exception as e:
        raise DataValidationError(f"Couldn't read that Excel file: {e}")
    if raw.empty:
        raise DataValidationError("The uploaded file has no rows.")
    return _normalize(raw)


def generate_excel_template() -> bytes:
    """Builds a starter .xlsx: a clean 'Products' sheet (headers + one example
    row) plus a separate 'Instructions' sheet, so the legend text never ends
    up parsed as data rows regardless of how many rows the user adds."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = REQUIRED_COLUMNS + OPTIONAL_COLUMNS
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    example = ["Basmati Rice 5kg", "Grocery", 480, 390, 42, 15, "", 5]
    example_fill = PatternFill("solid", fgColor="FFF9DB")
    for col_idx, v in enumerate(example, start=1):
        cell = ws.cell(row=2, column=col_idx, value=v)
        cell.fill = example_fill
        cell.font = Font(name="Arial")

    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = max(14, len(h) + 4)

    instructions = wb.create_sheet("Instructions")
    instructions.cell(row=1, column=1, value="How to fill in this template").font = Font(name="Arial", bold=True, size=13)
    notes = [
        "Row 2 (yellow) on the Products sheet is an example — replace or delete it, then add your own rows below it.",
        "Required columns: " + ", ".join(REQUIRED_COLUMNS),
        "Optional columns: " + ", ".join(OPTIONAL_COLUMNS),
        "expiry_days: leave blank for non-perishable items (rice, oil, household goods).",
        "expiry_days: number of days from today until the product expires (e.g. 5 for milk).",
        "price, cost, stock, reorder_level must be plain numbers — no currency symbols or commas.",
        "gst_rate: the GST slab as a plain number (e.g. 5, 12, 18) — leave blank for 0%.",
        "Do not add extra rows or notes on the Products sheet itself — every row there is read as a product.",
    ]
    for i, note in enumerate(notes, start=3):
        instructions.cell(row=i, column=1, value=f"- {note}").font = Font(name="Arial", size=10)
    instructions.column_dimensions["A"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------
# Database
# --------------------------------------------------------------------------
def list_sqlite_tables(path: str) -> list[str]:
    import sqlite3
    con = sqlite3.connect(path)
    try:
        rows = con.execute("SELECT name FROM sqlite_master WHERE type='table';").fetchall()
        return [r[0] for r in rows]
    finally:
        con.close()


def load_products_from_sqlite(path: str, table: str) -> pd.DataFrame:
    import sqlite3
    con = sqlite3.connect(path)
    try:
        raw = pd.read_sql(f"SELECT * FROM {table}", con)
    except Exception as e:
        raise DataValidationError(f"Couldn't read table '{table}': {e}")
    finally:
        con.close()
    if raw.empty:
        raise DataValidationError(f"Table '{table}' has no rows.")
    return _normalize(raw)


def test_db_connection(connection_string: str) -> tuple[bool, str]:
    """Returns (ok, message). Doesn't raise, so the UI can show a friendly result."""
    from sqlalchemy import create_engine, text
    try:
        engine = create_engine(connection_string)
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return True, "Connection successful."
    except Exception as e:
        return False, str(e)


def list_db_tables(connection_string: str) -> list[str]:
    from sqlalchemy import create_engine, inspect
    engine = create_engine(connection_string)
    return inspect(engine).get_table_names()


def load_products_from_db(connection_string: str, table_or_query: str, is_query: bool = False) -> pd.DataFrame:
    """table_or_query: a table name, or a full SELECT statement if is_query=True."""
    from sqlalchemy import create_engine, text
    try:
        engine = create_engine(connection_string)
        sql = table_or_query if is_query else f"SELECT * FROM {table_or_query}"
        with engine.connect() as conn:
            raw = pd.read_sql(text(sql) if is_query else sql, conn)
    except Exception as e:
        raise DataValidationError(f"Couldn't fetch data: {e}")
    if raw.empty:
        raise DataValidationError("The query returned no rows.")
    return _normalize(raw)
